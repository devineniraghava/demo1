# -*- coding: utf-8 -*-
"""
Created on Wed Apr 28 15:04:54 2021

@author: Devineni
"""




import pandas as pd
import os
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime


import pymysql
from sqlalchemy import create_engine

def prRed(skk): print("\033[31;1;m {}\033[00m" .format(skk)) 
schema = "cbo_winter"
engine = create_engine("mysql+pymysql://root:Password123@localhost/{}".format(schema),pool_pre_ping=True)
names = pd.read_sql_query('SHOW TABLES FROM {}'.format(schema), engine)

#%%
import datetime
import matplotlib.dates as mdates

import matplotlib.units as munits
from pylab import rcParams
rcParams['figure.figsize'] = 5.9111111111111105,3.8 # word
# rcParams['figure.figsize'] = 19,15 # word

plt.rcParams["font.family"] = "calibri"
plt.rcParams["font.weight"] = "normal"
plt.rcParams["font.size"] = 10






#%%


#%%
# df = pd.read_excel("sample.xlsx").set_index("sensor")

# column_names = list(df.columns)
# row_names = list(df.index.values)

# fig = plt.figure()
# ax = Axes3D(fig)

# lx= df.shape[1]       # Work out matrix dimensions
# ly= df.shape[0]
# xpos = np.arange(0,lx,1)    # Set up a mesh of positions
# ypos = np.arange(0,ly,1)
# xpos, ypos = np.meshgrid(xpos+0.5, ypos+0.5)


# xpos = xpos.flatten()   # Convert positions to 1D array
# ypos = ypos.flatten()
# zpos = np.zeros(lx*ly)

# dx = 0.2 * np.ones_like(zpos)
# dy = dx.copy()
# dz = df.values.flatten()

# cs = ['r', 'g', 'b', 'y'] * ly
# cs = ['#179C7D','#F29400','#1F82C0','#B1C800'] * ly

# ax.bar3d(xpos,ypos,zpos, dx, dy, dz,color=cs)

# #sh()
# ax.w_xaxis.set_ticklabels(column_names)
# ax.w_yaxis.set_ticklabels(row_names)
# ax.set_xlabel('Rooms')
# ax.set_ylabel('Sensor positions')
# ax.set_zlabel('Local Air exchange Index')
# ax.text2D(0.05, 0.75, "S_Sp5_ExOn", transform=ax.transAxes)

# ticksx = np.arange(0.5, 4, 1)
# plt.xticks(ticksx, column_names)

# ticksy = np.arange(0.6, 6, 1)
# plt.yticks(ticksy, row_names)

# plt.show()

# x = ['#179C7D','#F29400','#1F82C0','#E2001A','#B1C800']

# plt.savefig("3d_plot.png", dpi = 300)




#%% CBO
from easygui import *

from openpyxl import load_workbook
wb2 = load_workbook('C:/Users/Devineni/OneDrive - bwedu/MA_Raghavakrishna/1_Evaluation/results/df_eps.xlsx')

names = wb2.sheetnames

# msg ="Please select a sheet"
# title = "sheet selection"
# choices = names
experiment = "W_H_e0_Her"

#%%
df = pd.read_excel("C:/Users/Devineni/OneDrive - bwedu/4_Recirculation/python_files/demo1/Copy of df_eps.xlsx", sheet_name=experiment)

matches = ["1a_testo","1b","1c","1d","1e","1t","1l","2a_testo","2b","2c","2d","2e","2t","3a_testo","3b","3c","3d","3e","3t","4b","4c","4t"]
for i in matches:
    df.loc[df['index'].str.contains(i), "sensor"] = i




df.loc[df['eps_ai_n']>=1.16, 'color'] = '#29A500' 
df.loc[(df['eps_ai_n']<1.16) & (df['eps_ai_n']>1.10), 'color'] = '#00FF3F' 
df.loc[(df['eps_ai_n']<=1.10) & (df['eps_ai_n']>1.05), 'color'] = '#BFFF00' 
df.loc[(df['eps_ai_n']<=1.05) & (df['eps_ai_n']>0.94), 'color'] = '#FFFF00' 
df.loc[(df['eps_ai_n']<=0.94) & (df['eps_ai_n']>0.89), 'color'] = '#FFBF00' 
df.loc[(df['eps_ai_n']<=0.89) & (df['eps_ai_n']>0.84), 'color'] = '#FF3F00' 
df.loc[(df['eps_ai_n']<=0.84), 'color'] = '#A50052' 

df = df.loc[:,['eps_ai_n','color', 'sensor']].set_index("sensor")

df1 = pd.read_excel("C:/Users/Devineni/OneDrive - bwedu/4_Recirculation/python_files/demo1/Copy of df_eps.xlsx", sheet_name="Sheet1").set_index("sensor")

df3 = pd.concat([df,df1], axis = 1).dropna().reset_index()




colors = ["#29A500", "#00FF3F", "#BFFF00", "#FFFF00", "#FFBF00", "#FF3F00", "#A50052"]


#%% CBO

from matplotlib.lines import Line2D

import matplotlib.pyplot as plt
from mpl_toolkits.axes_grid1 import make_axes_locatable
img = plt.imread("CBO_6.1.png")
fig, ax = plt.subplots()
ax.imshow(img, extent=[0, 11, 0, 11])



# libraries
import matplotlib.pyplot as plt
import numpy as np
 
# create data
x = df3["x"].values.flatten()
y = df3["y"].values.flatten()
z = df3["z"].values.flatten()
# Change color with c and alpha
c= df3["color"].values.flatten()


scatter = ax.scatter(x, y, c=c, s=z*120, alpha=0.8)
legend_elements = [Line2D([0], [0], marker='o', color='w', label='1.16 und mehr',markerfacecolor=colors[0], markersize=10),
                    Line2D([0], [0], marker='o', color='w', label='1.11 - 1.15',markerfacecolor=colors[1], markersize=10),
                    Line2D([0], [0], marker='o', color='w', label='1.06 - 1.10',markerfacecolor=colors[2], markersize=10),
                    Line2D([0], [0], marker='o', color='w', label='0.95 - 1.05',markerfacecolor=colors[3], markersize=10),
                    Line2D([0], [0], marker='o', color='w', label='0.90 - 0.94',markerfacecolor=colors[4], markersize=10),
                    Line2D([0], [0], marker='o', color='w', label='0.85 - 0.89',markerfacecolor=colors[5], markersize=10),
                    Line2D([0], [0], marker='o', color='w', label='0.84 oder weniger',markerfacecolor=colors[6], markersize=10)]
ax.legend(handles=legend_elements, loc='upper right', bbox_to_anchor=(1.35, 0.9), labelspacing = 1, fontsize = 6)
ax.axis('off')
# ax.set_title("Local air exchange efficiency " +  experiment, fontsize = 6)


for line in range(0,df3.shape[0]):
      ax.text(df3.x[line], (df3.y[line]-0.1), round(df3.eps_ai_n[line],2), horizontalalignment='center', size=6, color='black', weight='normal')

# ax.add_artist(legend1)
plt.show()
path = "C:/Users/Devineni/OneDrive - bwedu/MA_Raghavakrishna/1_Evaluation/python_files/plots/cbo/"
plt.text(7.5, 4.1, r'$\mathrm{\epsilon^{a} = 0,56}$', fontsize=7)
plt.savefig("eps_ai.png", dpi = 600, bbox_inches='tight')  

